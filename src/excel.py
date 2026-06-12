"""Registro de proformas confirmadas en el Excel definitivo para Hacienda.

Toda la lógica del Excel vive aquí (modular, aislada del resto de la app).
Patrón robusto tomado del factura-processor: copia de seguridad antes de
escribir, reintentos si el Excel está abierto en el PC (LibreOffice/Excel por
Samba) y una cola de pendientes que se vacía sola en cuanto se libera el fichero.

Mapa de columnas (16) -> ver `HEADERS`. Las celdas calculadas se escriben como
fórmulas Excel vivas (Total = Base + IVA, etc.); el IVA es una fórmula cuando la
proforma tiene un único tipo, y un valor ya calculado si mezcla tipos.
"""

import os
import json
import time
import shutil
import fcntl
from contextlib import contextmanager
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from db import get_db

# ── Configuración (por entorno; defaults en el NAS, nunca en el rootfs del CT) ──
EXCEL_PATH = os.environ.get('EXCEL_PATH', '/mnt/empresa/facturas-emitidas.xlsx')
EXCEL_BACKUP_DIR = os.environ.get('EXCEL_BACKUP_DIR', '/mnt/empresa/backups-proformas')
EXCEL_PENDING_FILE = os.environ.get('EXCEL_PENDING_FILE', '/mnt/empresa/proforma-admin-pending-excel.json')
EXCEL_LOCK_FILE = os.environ.get('EXCEL_LOCK_FILE', '/mnt/empresa/.proforma-excel.lock')

MAX_RETRIES = 3
RETRY_DELAY = 2          # segundos entre reintentos si el Excel está bloqueado
BACKUP_RETENTION_DAYS = 30
MONEY_FORMAT = '#,##0.00" €"'

# Cabecera (orden real confirmado con el Excel de la empresa).
HEADERS = [
    'Índice', 'Fecha Factura', 'Nº Proforma', 'NIF/CIF', 'Tr', 'Agencia',
    'Provincia', 'Base', 'IVA', 'Total', 'Suplidos', 'Total + suplidos',
    'Cobrado', 'Comentarios', 'Nº Factura', 'Guía',
]
COL_WIDTHS = {
    'A': 7, 'B': 12, 'C': 15, 'D': 12, 'E': 5, 'F': 26, 'G': 14, 'H': 11,
    'I': 11, 'J': 11, 'K': 11, 'L': 15, 'M': 10, 'N': 34, 'O': 12, 'P': 18,
}

# Resultados posibles de un registro.
OK = 'ok'                      # fila escrita correctamente
EN_COLA = 'en_cola'            # Excel bloqueado -> encolada para reintentar
YA_REGISTRADA = 'ya_registrada'
NO_EXISTE = 'no_existe'
_BLOQUEADO = 'bloqueado'       # interno: el save falló por PermissionError


# ── Helpers de bajo nivel ─────────────────────────────────────────────────────

def _ensure_parent(path):
    carpeta = os.path.dirname(path)
    if carpeta:
        os.makedirs(carpeta, exist_ok=True)


@contextmanager
def _file_lock():
    """Serializa las escrituras de este proceso sobre el Excel."""
    _ensure_parent(EXCEL_LOCK_FILE)
    f = open(EXCEL_LOCK_FILE, 'w')
    try:
        fcntl.flock(f, fcntl.LOCK_EX)
        yield
    finally:
        fcntl.flock(f, fcntl.LOCK_UN)
        f.close()


def _load_or_create_workbook():
    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
        return wb, wb.active
    wb = Workbook()
    ws = wb.active
    ws.title = 'Facturas emitidas'
    ws.append(HEADERS)
    for col in range(1, len(HEADERS) + 1):
        ws.cell(1, col).font = Font(bold=True)
    for letra, ancho in COL_WIDTHS.items():
        ws.column_dimensions[letra].width = ancho
    ws.freeze_panes = 'A2'
    return wb, ws


def _backup():
    """Copia con sello de tiempo antes de tocar el Excel + rotación a 30 días."""
    if not os.path.exists(EXCEL_PATH):
        return
    os.makedirs(EXCEL_BACKUP_DIR, exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d-%H%M%S')
    shutil.copy2(EXCEL_PATH, os.path.join(EXCEL_BACKUP_DIR, f'facturas-emitidas_{ts}.xlsx'))
    limite = time.time() - BACKUP_RETENTION_DAYS * 86400
    for nombre in os.listdir(EXCEL_BACKUP_DIR):
        if nombre.startswith('facturas-emitidas_') and nombre.endswith('.xlsx'):
            ruta = os.path.join(EXCEL_BACKUP_DIR, nombre)
            if os.path.getmtime(ruta) < limite:
                try:
                    os.remove(ruta)
                except OSError:
                    pass


def _save_con_reintentos(wb):
    """Guarda reintentando si el Excel está abierto. True si lo consigue."""
    _ensure_parent(EXCEL_PATH)
    for intento in range(MAX_RETRIES):
        try:
            wb.save(EXCEL_PATH)
            return True
        except PermissionError:
            if intento < MAX_RETRIES - 1:
                time.sleep(RETRY_DELAY)
    return False


# ── Construcción y escritura de la fila ───────────────────────────────────────

def _build_row(conn, proforma):
    """Reúne los datos de proforma + cliente + guías para una fila del Excel."""
    cliente = None
    if proforma['cliente_id']:
        cliente = conn.execute(
            "SELECT * FROM clientes WHERE id = ?", (proforma['cliente_id'],)
        ).fetchone()

    guias = conn.execute(
        """SELECT g.nombre FROM guias g
           JOIN proforma_guias pg ON g.id = pg.guia_id
           WHERE pg.proforma_id = ?
           ORDER BY g.nombre""",
        (proforma['id'],)
    ).fetchall()

    tipos = {
        round(l['porcentaje_iva'], 4)
        for l in conn.execute(
            "SELECT porcentaje_iva FROM proforma_lineas WHERE proforma_id = ?",
            (proforma['id'],)
        ).fetchall()
    }

    from datetime import date as _date
    fecha_raw = proforma['fecha']
    try:
        fecha_obj = _date.fromisoformat(fecha_raw) if fecha_raw else None
    except (ValueError, TypeError):
        fecha_obj = None

    return {
        'numero': proforma['numero_proforma'],
        'fecha': fecha_obj,
        'nif': (cliente['nif_cif'] if cliente else '') or '',
        'trimestre': proforma['trimestre'],
        'agencia': (cliente['nombre_agencia'] if cliente else '') or '',
        'provincia': (cliente['provincia'] if cliente else '') or '',
        'base': round(proforma['base'] or 0, 2),
        'iva_total': round(proforma['iva_total'] or 0, 2),
        'suplidos': round(proforma['suplidos'] or 0, 2),
        'comentarios': proforma['comentarios'] or '',
        'guias': ', '.join(g['nombre'] for g in guias),
        'iva_uniforme': len(tipos) == 1,
        'tipo_iva': next(iter(tipos)) if len(tipos) == 1 else None,
    }


def _write_row(ws, n, row):
    """Escribe la fila n con valores, fórmulas y formato de moneda."""
    ws.cell(n, 1, n - 1)               # A  Índice (correlativo de datos)
    if row.get('fecha'):
        ws.cell(n, 2, row['fecha']).number_format = 'DD/MM/YYYY'  # B  Fecha
    ws.cell(n, 3, row['numero'])       # C  Nº Proforma
    ws.cell(n, 4, row['nif'])          # D  NIF/CIF
    ws.cell(n, 5, row['trimestre'])    # E  Tr
    ws.cell(n, 6, row['agencia'])      # F  Agencia
    ws.cell(n, 7, row['provincia'])    # G  Provincia
    ws.cell(n, 8, row['base']).number_format = MONEY_FORMAT          # H  Base

    # I  IVA: fórmula si el tipo es único; valor calculado si mezcla tipos.
    if row['iva_uniforme'] and row['tipo_iva']:
        ws.cell(n, 9, f"=H{n}*{row['tipo_iva'] / 100:g}")
    else:
        ws.cell(n, 9, row['iva_total'])
    ws.cell(n, 9).number_format = MONEY_FORMAT

    ws.cell(n, 10, f"=H{n}+I{n}").number_format = MONEY_FORMAT       # J  Total

    # K  Suplidos: solo si la proforma trae alguno; si no, se deja para rellenar a mano.
    if row['suplidos'] > 0:
        ws.cell(n, 11, row['suplidos']).number_format = MONEY_FORMAT

    ws.cell(n, 12, f"=J{n}+K{n}").number_format = MONEY_FORMAT       # L  Total + suplidos
    # M  Cobrado -> manual
    ws.cell(n, 14, row['comentarios'])  # N  Comentarios
    # O  Nº Factura -> manual (Factusol)
    ws.cell(n, 16, row['guias'])        # P  Guía


# ── Cola de pendientes (cuando el Excel está abierto) ─────────────────────────

def _cargar_cola():
    if not os.path.exists(EXCEL_PENDING_FILE):
        return []
    try:
        with open(EXCEL_PENDING_FILE) as f:
            datos = json.load(f)
        return datos if isinstance(datos, list) else []
    except (json.JSONDecodeError, OSError):
        return []


def _guardar_cola(ids):
    _ensure_parent(EXCEL_PENDING_FILE)
    with open(EXCEL_PENDING_FILE, 'w') as f:
        json.dump(ids, f)


def _encolar(proforma_id):
    ids = _cargar_cola()
    if proforma_id not in ids:
        ids.append(proforma_id)
        _guardar_cola(ids)


def contar_pendientes():
    return len(_cargar_cola())


# ── API pública ───────────────────────────────────────────────────────────────

def _intentar_escribir(proforma_id):
    """Intenta escribir la fila. Devuelve OK / YA_REGISTRADA / NO_EXISTE / _BLOQUEADO."""
    with get_db() as conn:
        proforma = conn.execute(
            "SELECT * FROM proformas WHERE id = ?", (proforma_id,)
        ).fetchone()
        if proforma is None:
            return NO_EXISTE
        if proforma['exportada_excel']:
            return YA_REGISTRADA
        row = _build_row(conn, proforma)

    with _file_lock():
        _backup()
        wb, ws = _load_or_create_workbook()
        _write_row(ws, ws.max_row + 1, row)
        if not _save_con_reintentos(wb):
            return _BLOQUEADO

    with get_db() as conn:
        conn.execute(
            "UPDATE proformas SET exportada_excel = 1 WHERE id = ?", (proforma_id,)
        )
    return OK


def registrar_proforma(proforma_id):
    """Registra una proforma en el Excel. Si está bloqueado, la encola."""
    resultado = _intentar_escribir(proforma_id)
    if resultado == _BLOQUEADO:
        _encolar(proforma_id)
        return EN_COLA
    return resultado


def _buscar_fila(ws, numero_proforma):
    """Devuelve el número de fila (1-based) cuya columna C = Nº Proforma, o None."""
    for row in ws.iter_rows(min_row=2):
        if row[2].value == numero_proforma:  # columna C = Nº Proforma
            return row[0].row
    return None


def marcar_cobrado_excel(numero_proforma, fecha_cobro):
    """Escribe la fecha de cobro en la columna 'Cobrado' (M) de su fila.

    `fecha_cobro` es una fecha ISO (str) o un date. Devuelve True si se escribió,
    False si no se encontró la fila o no existe el fichero, None si estaba bloqueado.
    """
    if not os.path.exists(EXCEL_PATH):
        return False
    from datetime import date as _date
    try:
        fecha_obj = _date.fromisoformat(fecha_cobro) if isinstance(fecha_cobro, str) else fecha_cobro
    except (ValueError, TypeError):
        fecha_obj = None
    with _file_lock():
        _backup()
        wb, ws = _load_or_create_workbook()
        fila = _buscar_fila(ws, numero_proforma)
        if fila is None:
            return False
        celda = ws.cell(fila, 13, fecha_obj)   # M  Cobrado
        if fecha_obj is not None:
            celda.number_format = 'DD/MM/YYYY'
        if not _save_con_reintentos(wb):
            return None  # Excel bloqueado
    return True


def desmarcar_cobrado_excel(numero_proforma):
    """Vacía la columna 'Cobrado' (M) de la fila indicada (deshacer cobro).

    Mismos valores de retorno que marcar_cobrado_excel.
    """
    if not os.path.exists(EXCEL_PATH):
        return False
    with _file_lock():
        _backup()
        wb, ws = _load_or_create_workbook()
        fila = _buscar_fila(ws, numero_proforma)
        if fila is None:
            return False
        ws.cell(fila, 13).value = None   # M  Cobrado (cell(...,None) sería no-op en openpyxl)
        if not _save_con_reintentos(wb):
            return None  # Excel bloqueado
    return True


def eliminar_fila_excel(numero_proforma):
    """Elimina del Excel la fila del número de proforma indicado.

    Devuelve True si se eliminó, False si no se encontró o el fichero no existe,
    None si el fichero estaba bloqueado.
    """
    if not os.path.exists(EXCEL_PATH):
        return False
    with _file_lock():
        _backup()
        wb, ws = _load_or_create_workbook()
        fila_a_borrar = _buscar_fila(ws, numero_proforma)
        if fila_a_borrar is None:
            return False
        ws.delete_rows(fila_a_borrar)
        # Re-numerar índice (columna A) de todas las filas de datos
        for i, row_cells in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=1):
            row_cells[0].value = i
        if not _save_con_reintentos(wb):
            return None  # Excel bloqueado
    return True


def drain_pending():
    """Vacía la cola de pendientes. Devuelve cuántas filas se han escrito."""
    ids = _cargar_cola()
    if not ids:
        return 0
    restantes, escritas = [], 0
    for pid in ids:
        resultado = _intentar_escribir(pid)
        if resultado == _BLOQUEADO:
            restantes.append(pid)        # sigue bloqueado: se reintenta más tarde
        elif resultado == OK:
            escritas += 1
        # YA_REGISTRADA / NO_EXISTE -> se descarta de la cola
    _guardar_cola(restantes)
    return escritas
